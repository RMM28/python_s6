import tkinter as tk
from tkinter import messagebox
import openpyxl
from tkinter import ttk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Expense Tracker")
        self.expense_file = "expenses.xlsx"
        self.expenses = []

        # Month entry
        self.label_month = tk.Label(root, text="Month:")
        self.label_month.grid(row=0, column=0, padx=5, pady=5)

        self.entry_month = tk.Entry(root)
        self.entry_month.grid(row=0, column=1, padx=5, pady=5)

        # Expense entry
        self.label_expense = tk.Label(root, text="Item:")
        self.label_expense.grid(row=1, column=0, padx=5, pady=5)

        self.entry_expense = tk.Entry(root)
        self.entry_expense.grid(row=1, column=1, padx=5, pady=5)

        # Amount entry
        self.label_amount = tk.Label(root, text="Amount:")
        self.label_amount.grid(row=2, column=0, padx=5, pady=5)

        self.entry_amount = tk.Entry(root)
        self.entry_amount.grid(row=2, column=1, padx=5, pady=5)

        # Buttons
        self.button_add = tk.Button(root, text="Add Item", command=self.add_expense)
        self.button_add.grid(row=2, column=0, columnspan=2, padx=5, pady=5, sticky="we")

        self.button_graph = tk.Button(root, text="Show Graph", command=self.show_graph)
        self.button_graph.grid(row=3, column=0, columnspan=2, padx=5, pady=5, sticky="we")

        # Listbox
        self.listbox_expenses = tk.Listbox(root, width=40)
        self.listbox_expenses.grid(row=5, column=0, columnspan=2, padx=5, pady=5)

        self.load_expenses_from_file()

    def add_expense(self):
        month = self.entry_month.get().strip()
        expense = self.entry_expense.get().strip()
        amount = self.entry_amount.get().strip()
        if month and expense and amount:
            try:
                amount = float(amount)  # Ensure amount is a valid number
                self.expenses.append((month, expense, amount))
                self.update_expense_listbox()
                self.save_expenses_to_file()
                self.entry_month.delete(0, tk.END)
                self.entry_expense.delete(0, tk.END)
                self.entry_amount.delete(0, tk.END)
            except ValueError:
                messagebox.showwarning("Warning", "Please enter a valid number for amount.")
        else:
            messagebox.showwarning("Warning", "Please enter month, expense and amount.")

    def update_expense_listbox(self):
        self.listbox_expenses.delete(0, tk.END)
        for month, expense, amount in self.expenses:
            self.listbox_expenses.insert(tk.END, f"{month} - {expense}: ₹ {amount}")

    def load_expenses_from_file(self):
        try:
            workbook = openpyxl.load_workbook(self.expense_file)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                self.expenses.append((row[0], row[1], row[2]))
            workbook.close()
            self.update_expense_listbox()
        except FileNotFoundError:
            pass

    def save_expenses_to_file(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for month, expense, amount in self.expenses:
            sheet.append([month, expense, amount])
        workbook.save(self.expense_file)

    def show_graph(self):
        # Prepare data for plotting
        monthly_sums = {}
        for month, _, amount in self.expenses:
            if month in monthly_sums:
                monthly_sums[month] += amount
            else:
                monthly_sums[month] = amount

        months = list(monthly_sums.keys())
        totals = list(monthly_sums.values())

        # Plotting
        fig, ax = plt.subplots()
        ax.bar(months, totals, color='blue')
        ax.set_title('Monthly Expenses')
        ax.set_xlabel('Month')
        ax.set_ylabel('Total Expenses')

        # Embedding matplotlib graph in Tkinter window
        canvas = FigureCanvasTkAgg(fig, master=self.root)
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.grid(row=6, column=0, columnspan=2, pady=5)
        canvas.draw()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseTracker(root)
    root.mainloop()
